#!/usr/bin/env python3
"""Generate task test-case Excel for Mission task IDs 40000–40018."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MISSION_XLSX = Path(r"d:\WorkSpacelml\Branches\RCT\Excel\Mission.xlsx")
OUTPUT_XLSX = Path(r"C:\Users\user\Desktop\任务测试用例_40000-40018.xlsx")

TASK_ID_MIN = 40000
TASK_ID_MAX = 40018

HEADERS = [
    "任务ID",
    "任务名称",
    "阶段",
    "进局前准备（PreCondition）",
    "局内/局外",
    "局内检查点",
    "局外检查点",
    "预期结果",
    "备注",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATA_FILL = PatternFill("solid", fgColor="D9E8F7")
WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet_rows(wb: openpyxl.Workbook, name: str, data_start_row: int = 4):
    ws = wb[name]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(header) if h is not None}
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or row[0] is None:
            continue
        yield row, idx


def load_condition_type_names(wb: openpyxl.Workbook) -> dict[int, str]:
    ws = wb["@Condition枚举表"]
    mapping: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        label = str(row[0]).strip()
        try:
            type_id = int(row[1])
        except (TypeError, ValueError):
            continue
        if label and label != "说明":
            mapping[type_id] = label
    return mapping


def load_conditions(wb: openpyxl.Workbook) -> dict[int, dict]:
    conditions: dict[int, dict] = {}
    for row, idx in _sheet_rows(wb, "Condition"):
        try:
            cid = int(row[0])
        except (TypeError, ValueError):
            continue
        conditions[cid] = {
            "type": row[idx.get("条件类型", 7)],
            "value": row[idx.get("条件值", 8)],
            "ingame": bool(row[idx.get("是否为局内任务", 6)]),
            "text_key": row[idx.get("任务条件文本", 2)],
        }
    return conditions


def load_tasks(wb: openpyxl.Workbook) -> list[dict]:
    tasks: list[dict] = []
    for row, idx in _sheet_rows(wb, "Task"):
        try:
            tid = int(row[0])
        except (TypeError, ValueError):
            continue
        if tid < TASK_ID_MIN or tid > TASK_ID_MAX:
            continue

        raw_conds = str(row[idx.get("任务目标多个目标分号隔开", 21)] or "")
        cond_ids = [int(x.strip()) for x in raw_conds.split(";") if x.strip()]

        pre = row[idx.get("前置任务", 17)]
        pre_task = int(pre) if pre not in (None, "", "null") else None

        auto_raw = row[idx.get("是否自动接取", 14)]
        auto_accept = str(auto_raw).strip() in ("1", "True", "true", "是")

        tasks.append(
            {
                "id": tid,
                "name": f"营地任务 {tid}",
                "unlock_type": row[idx.get("解锁条件", 15)],
                "unlock_param": row[idx.get("解锁参数", 16)],
                "pre_task": pre_task,
                "cond_ids": cond_ids,
                "auto_accept": auto_accept,
            }
        )
    tasks.sort(key=lambda t: t["id"])
    return tasks


def format_condition_line(cid: int, cond: dict, type_names: dict[int, str]) -> str:
    type_id = cond.get("type")
    type_name = type_names.get(int(type_id), f"类型{type_id}") if type_id is not None else "未知类型"
    value = cond.get("value", "")
    return f"条件{cid}【{type_name}】目标值={value}"


def build_accept_row(task: dict) -> list:
    checkpoint = (
        f"任务列表能看到本任务，标题和描述正确；任务ID={task['id']}"
        f"；解锁条件={task['unlock_type']}，参数={task['unlock_param']}"
    )
    if task["pre_task"] is not None:
        checkpoint += f"；前置任务={task['pre_task']}"

    expected = "任务变为【进行中】，目标描述和 PreCondition 提示正常"
    if task["auto_accept"]:
        expected += "（或已自动接取）"

    remark = f"自动接取={'是' if task['auto_accept'] else '否'}"
    return [
        task["id"],
        task["name"],
        "接取",
        "",
        "局外",
        "",
        checkpoint,
        expected,
        remark,
    ]


def build_execute_row(task: dict, conditions: dict[int, dict], type_names: dict[int, str]) -> list:
    lines: list[str] = []
    any_ingame = False
    for cid in task["cond_ids"]:
        cond = conditions.get(cid)
        if not cond:
            lines.append(f"条件{cid}【配置缺失】")
            continue
        if cond["ingame"]:
            any_ingame = True
        lines.append(format_condition_line(cid, cond, type_names))

    checkpoint_text = "\n".join(lines)
    in_out = "局内" if any_ingame else "局外"
    if in_out == "局内":
        in_checkpoint = checkpoint_text
        out_checkpoint = ""
        expected = "进度条/计数与 Condition 配置一致，局内任务在局内 HUD 可见"
    else:
        in_checkpoint = ""
        out_checkpoint = checkpoint_text
        expected = "进度条/计数与 Condition 配置一致"

    remark = f"自动接取={'是' if task['auto_accept'] else '否'}"
    return [
        task["id"],
        task["name"],
        "执行",
        "",
        in_out,
        in_checkpoint,
        out_checkpoint,
        expected,
        remark,
    ]


def build_complete_row(task: dict) -> list:
    remark = f"自动接取={'是' if task['auto_accept'] else '否'}"
    return [
        task["id"],
        task["name"],
        "完成",
        "",
        "局外",
        "",
        "奖励到账；任务变【已完成】；下一条前置任务解锁",
        "领奖成功无报错，可继续下一条顺序号",
        remark,
    ]


def build_rows(tasks: list[dict], conditions: dict[int, dict], type_names: dict[int, str]) -> list[list]:
    rows: list[list] = []
    for task in tasks:
        rows.append(build_accept_row(task))
        rows.append(build_execute_row(task, conditions, type_names))
        rows.append(build_complete_row(task))
    return rows


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(HEADERS[col_idx - 1]))
        for cell in ws[letter]:
            if cell.value is None:
                continue
            for part in str(cell.value).split("\n"):
                max_len = max(max_len, len(part))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def write_workbook(rows: list[list], output: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务测试用例"

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = DATA_FILL
            cell.alignment = WRAP

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    if not MISSION_XLSX.is_file():
        raise SystemExit(f"Mission.xlsx not found: {MISSION_XLSX}")

    wb = openpyxl.load_workbook(MISSION_XLSX, read_only=True, data_only=True)
    type_names = load_condition_type_names(wb)
    conditions = load_conditions(wb)
    tasks = load_tasks(wb)
    wb.close()

    if len(tasks) != TASK_ID_MAX - TASK_ID_MIN + 1:
        found = [t["id"] for t in tasks]
        missing = [i for i in range(TASK_ID_MIN, TASK_ID_MAX + 1) if i not in found]
        raise SystemExit(f"Expected 19 tasks, got {len(tasks)}. Missing: {missing}")

    rows = build_rows(tasks, conditions, type_names)
    if len(rows) != 57:
        raise SystemExit(f"Expected 57 rows, got {len(rows)}")

    write_workbook(rows, OUTPUT_XLSX)
    print(f"Wrote {len(rows)} rows to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
