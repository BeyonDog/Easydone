# -*- coding: utf-8 -*-
"""Export task test cases with 2.3 rules, i18n audit, cross-ref checks, config issue sheet."""
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RCT_ROOT = Path(r"D:\WorkSpacelml\Branches\RCT")
MISSION_XLSX = RCT_ROOT / "Excel" / "Mission.xlsx"
EXCEL_ROOT = RCT_ROOT / "Excel"
LANG_ROOT = RCT_ROOT / "Client" / "Assets" / "AssetBox" / "ResFolder" / "UI" / "Loc" / "Lang"
CONFIG_PATH = Path(r"C:\Users\user\AppData\Roaming\com.easydone.desktop\config.json")
OUT_PATH = Path.home() / "Desktop" / "任务全量用例-老任务首胜BP-v3.xlsx"

CHAIN_NAMES = {
    "300": "任务链300（晨曦圣所主线）", "301": "任务链301", "302": "任务链302",
    "303": "任务链303（战士）", "304": "任务链304（游侠）", "305": "任务链305（法师）",
    "306": "任务链306（牧师）", "307": "任务链307（盗贼）", "308": "任务链308（召唤师）",
}

INGAME_CDT = {1, 2, 3, 4, 22, 23, 24, 27, 28, 29, 30, 31, 39, 40, 45}
FOREIGN_LANGS = ["en", "pt", "vi", "th", "id", "es", "ms"]
ALL_LANGS = ["zh-Hans"] + FOREIGN_LANGS
CODE_LANGS = ["zh-Hans", "en", "pt", "vi", "ja", "th", "id", "es", "ms"]

CDT_NAME = {
    1: "击杀", 2: "到达位置", 3: "带出物品", 4: "交互", 5: "登录", 6: "参与对局",
    7: "账号等级", 8: "带出价值", 9: "账号绑定", 10: "提交物品", 11: "商店出售金币",
    12: "交易行出售金币", 13: "出售金币", 14: "造成伤害", 15: "达到段位", 16: "收藏品总价值",
    17: "打造次数", 18: "商店购买次数", 19: "商店购买消耗金币", 20: "交易行上架",
    21: "交易行购买消耗金币", 22: "局内消耗物品", 23: "产生治疗量", 24: "组队共享击杀",
    26: "职业组解锁", 27: "成功撤离", 28: "进局携带技能", 29: "进局携带精通",
    30: "进局（对局计数）", 31: "进局携带物品", 32: "容器等级", 33: "收藏品进度累积",
    34: "收藏品套装完成", 35: "加入公会", 36: "谷歌评分", 37: "账户持有金币",
    38: "炼金炼制", 39: "完成局内紧急任务", 40: "连续胜利撤离", 41: "营地等级",
    42: "收藏品套装数量", 43: "小迅雷下载", 44: "猎人天赋节点", 45: "局内存活时长",
}

PRECDT_NAME = {
    1: "指定怪物", 2: "指定怪物类型", 3: "指定位置Trigger", 4: "指定物品ID", 5: "指定物品类型",
    6: "指定物品品质", 7: "指定交互物ID(已废弃)", 8: "指定交互行为", 9: "职业精通ID(2.3)",
    10: "指定怪物大类", 11: "指定宝箱类型", 12: "指定段位", 13: "指定收藏价值",
    14: "是否成功撤离", 15: "对局难度MatchMode", 16: "爆头击杀", 17: "击杀间距",
    18: "入局时间限制(秒)", 19: "装备分限制", 20: "进局力量", 21: "进局敏捷", 22: "进局智力",
    23: "进局体质", 24: "进局感知", 25: "进局携带某物品ID", 26: "指定技能ID", 27: "伤害类型",
    28: "怪物Tag(仅可配=)", 29: "交互物Tag", 30: "职业Tag", 31: "进局队伍人数",
    32: "打造配方ID(已废弃)", 33: "商店商品ID", 34: "大厅队伍状态GroupMode", 35: "带出物品价值",
    36: "打造配方子类型", 37: "带出价值", 38: "撤离地图LevelID", 39: "精通ID",
    40: "容器类型", 41: "收藏品套装ID(仅单个)", 42: "职业组ID(2.3解锁)", 43: "自身职业精通ID",
    44: "炼金配方ID", 45: "是否开启排位", 46: "物品子类型", 47: "紧急事件ID",
    48: "不使用药品", 49: "玩法GameMode", 50: "小迅雷下载Tag", 51: "猎人天赋节点ID",
    52: "职业组ID(2.3)", 53: "自身职业组ID(2.3)",
}

CMP_NAME = {0: "=", 1: ">=", 2: "<=", 3: "!="}
DEPRECATED_PRECDT = {
    7: "PreCdt=7(ITERACTIVE_ID)已废弃，不应再配",
    32: "PreCdt=32(MAKEITEM_MAKEID)已废弃，应改用PreCdt=36",
}
VALID_PRECDT = set(PRECDT_NAME)
COMMON_PRECDT = {14, 15, 31, 34, 45, 49}
CDT_PRECDT_FAMILY = {
    1: {1, 2, 9, 10, 16, 17, 18, 28, 30},
    2: {3},
    3: {4, 5, 6, 25, 35, 37, 46},
    4: {7, 8, 11, 29},
    10: {4, 5, 6, 46},
    14: {27},
    15: {12},
    16: {13, 41},
    17: {32, 36, 44},
    26: {42},
    28: {26},
    29: {9, 39, 43},
    31: {25},
    33: {41},
    34: {41},
    39: {47},
}
PRECDT_REF_TABLE = {
    1: ("Monster.xlsx", "monster"),
    2: ("Monster.xlsx", "monster_type"),
    4: ("Item.xlsx", "item"),
    9: ("Class.xlsx", "class"),
    10: ("Monster.xlsx", "monster_group"),
    11: ("Chest.xlsx", "chest"),
    12: ("Rank.xlsx", "rank"),
    13: ("Collection.xlsx", "collection_value"),
    25: ("Item.xlsx", "item"),
    33: ("Merchant_Normal.xlsx", "merchant"),
    36: ("MakeItem.xlsx", "make_subtype"),
    38: ("Level.xlsx", "level"),
    39: ("Class.xlsx", "mastery"),
    41: ("Collection.xlsx", "collection_set"),
    42: ("Class.xlsx", "class_group"),
    44: ("MakeItem.xlsx", "alchemy"),
    46: ("Item.xlsx", "item_subtype"),
    52: ("Class.xlsx", "class_group"),
    53: ("Class.xlsx", "class_group"),
}


def parse_int(v):
    if v is None or v == "" or str(v).lower() == "null":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def parse_string_key_table(path):
    text = path.read_text(encoding="utf-8")
    keys = {}
    for m in re.finditer(r"stringKey: (\S+)\s+intKey: (-?\d+)", text):
        keys[m.group(1)] = int(m.group(2))
    return keys


def parse_lang_asset(path):
    text = path.read_text(encoding="utf-8")
    trans = {}
    for m in re.finditer(r'translation: "(.*?)"\s+key: (-?\d+)', text, re.S):
        trans[int(m.group(2))] = m.group(1)
    return trans


def load_i18n():
    sk_path = LANG_ROOT / "StringKeyTable.asset"
    string_keys = parse_string_key_table(sk_path) if sk_path.exists() else {}
    lang_data = {}
    for lc in ALL_LANGS:
        p = LANG_ROOT / f"{lc}.asset"
        if p.exists():
            lang_data[lc] = parse_lang_asset(p)
    missing_assets = [lc for lc in CODE_LANGS if lc not in lang_data and lc != "ja"]
    if "ja" not in lang_data:
        missing_assets.append("ja")
    return string_keys, lang_data, missing_assets


def load_excel_keys(xlsx_path, sheet_name, min_row=6):
    import openpyxl
    ids = set()
    if not xlsx_path.exists():
        return ids
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return ids
    for row in wb[sheet_name].iter_rows(min_row=min_row, values_only=True):
        if row and row[0] is not None:
            v = parse_int(row[0])
            if v is not None:
                ids.add(v)
    wb.close()
    return ids


def load_ref_tables():
    # 为了尽量降低误报，尽可能做“存在性 + 基本内容”校验
    refs = {}

    refs["reward"] = load_excel_keys(EXCEL_ROOT / "Reward.xlsx", "Reward")

    # Dialog：Task 引用的是 DialogID（Dialog sheet 第一列）
    refs["dialog"] = load_excel_keys(EXCEL_ROOT / "Dialog.xlsx", "Dialog")
    dialog_content_has = set()
    import openpyxl

    if (EXCEL_ROOT / "Dialog.xlsx").exists():
        wb = openpyxl.load_workbook(EXCEL_ROOT / "Dialog.xlsx", read_only=True, data_only=True)
        if "DialogContent" in wb.sheetnames:
            for row in wb["DialogContent"].iter_rows(min_row=6, values_only=True):
                if not row or row[0] is None:
                    continue
                content_id = parse_int(row[0])
                if content_id is None:
                    continue
                # 经验规律：DialogContentID = DialogID * 10 + 序号
                dialog_id = content_id // 10
                dialog_content_has.add(dialog_id)
        wb.close()
    refs["dialog_has_content"] = dialog_content_has

    # Jump：Task/Condition 引用的是 JumpToPanel.JumpId（第一列）
    refs["jump_all"] = load_excel_keys(EXCEL_ROOT / "Jumptopanel.xlsx", "JumpToPanel")
    refs["jump_disabled"] = set()
    if (EXCEL_ROOT / "Jumptopanel.xlsx").exists():
        wb = openpyxl.load_workbook(EXCEL_ROOT / "Jumptopanel.xlsx", read_only=True, data_only=True)
        if "JumpToPanel" in wb.sheetnames:
            for row in wb["JumpToPanel"].iter_rows(min_row=6, values_only=True):
                if not row or row[0] is None:
                    continue
                jump_id = parse_int(row[0])
                disabled = parse_int(row[2])  # DisabledCondition 列
                if jump_id is None:
                    continue
                if disabled == 1:
                    refs["jump_disabled"].add(jump_id)
        wb.close()

    # 常用引用表：只做“ID 存在性”
    refs["monster"] = load_excel_keys(EXCEL_ROOT / "Monster.xlsx", "Monster")
    refs["item"] = load_excel_keys(EXCEL_ROOT / "Item.xlsx", "Item")
    refs["class"] = load_excel_keys(EXCEL_ROOT / "Class.xlsx", "Class")
    refs["rank"] = load_excel_keys(EXCEL_ROOT / "Rank.xlsx", "Rank")
    refs["collection"] = load_excel_keys(EXCEL_ROOT / "Collection.xlsx", "Collection")
    refs["level"] = load_excel_keys(EXCEL_ROOT / "Level.xlsx", "Level")
    refs["makeitem"] = load_excel_keys(EXCEL_ROOT / "MakeItem.xlsx", "MakeItem")
    refs["chest"] = load_excel_keys(EXCEL_ROOT / "Chest.xlsx", "Chest")
    refs["interactive"] = load_excel_keys(EXCEL_ROOT / "Interactive.xlsx", "Interactive")
    return refs


def split_semicolon_list(raw):
    if raw is None:
        return []
    return [x.strip() for x in str(raw).split(";") if x.strip() and x.strip().lower() != "null"]


def load_mission_tables():
    import openpyxl
    wb = openpyxl.load_workbook(MISSION_XLSX, read_only=True, data_only=True)

    patterns = {}
    for row in wb["ConditionPatternStringkey"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        patterns[parse_int(row[0])] = str(row[1] or "")

    task_pre_loc = {}
    for row in wb["TaskPreLoc"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        tid = str(parse_int(row[0]))
        task_pre_loc[tid] = {"title_key": str(row[1] or ""), "context_key": str(row[2] or "")}

    condition_pre_loc = {}
    for row in wb["ConditionPreLoc"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        cid = str(parse_int(row[0]))
        text_list = split_semicolon_list(row[2])
        color_list = split_semicolon_list(row[3])
        condition_pre_loc[cid] = {
            "pattern": parse_int(row[1]),
            "text_list": text_list,
            "color_list": color_list,
        }

    chains = {}
    for row in wb["TaskChain"].iter_rows(min_row=6, values_only=True):
        if not row or row[1] is None:
            continue
        chain_id = str(parse_int(row[1]))
        chains[chain_id] = {
            "name_key": str(row[3] or ""),
            "desc_key": str(row[4] or ""),
            "tab_id": parse_int(row[2]),
        }

    tabs = {}
    for row in wb["TaskTab"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        tabs[str(parse_int(row[0]))] = str(row[1] or "")

    chain_rewards = {}
    for row in wb["TaskChainReward"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        chain_id = str(parse_int(row[1]))
        chain_rewards.setdefault(chain_id, []).append({
            "reward_id": parse_int(row[3]),
            "task_count": parse_int(row[2]),
        })

    conds = {}
    for row in wb["Condition"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        cid = str(parse_int(row[0]))
        if not cid:
            continue
        pre = []
        for i in range(6):
            base = 9 + i * 3
            pt = parse_int(row[base])
            if pt in (None, 0):
                continue
            pre.append((pt, parse_int(row[base + 1]), row[base + 2]))
        conds[cid] = {
            "remark": row[1],
            "text_key": str(row[2] or ""),
            "goto": parse_int(row[3]),
            "ingame_flag": str(row[6]).strip().lower() == "true",
            "cdt": parse_int(row[7]),
            "cdt_val": row[8],
            "pre": pre,
            "relation": parse_int(row[27]),
        }

    tasks = {}
    for row in wb["Task"].iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        tid = str(parse_int(row[0]))
        if not tid:
            continue
        cids = split_semicolon_list(row[21])
        tasks[tid] = {
            "task_id": tid,
            "name": row[1],
            "desc": row[2],
            "title_key": str(row[3] or ""),
            "context_key": str(row[4] or ""),
            "type": parse_int(row[5]),
            "chain": str(parse_int(row[7]) or ""),
            "client_order": parse_int(row[9]),
            "start_time": str(row[10] or ""),
            "end_time": str(row[11] or ""),
            "condition_relation": parse_int(row[12]),
            "goto_group": str(row[13]).strip().lower() == "true",
            "auto_accept": str(row[14]).strip().lower() == "true",
            "unlock": str(row[15] or "").strip(),
            "unlock_param": str(row[16] or "").strip(),
            "pre_task": str(row[17] or "").strip(),
            "get_dialog": parse_int(row[18]),
            "claim_dialog": parse_int(row[19]),
            "image_path": str(row[20] or ""),
            "conds": cids,
            "level_obj_tag": split_semicolon_list(row[22]),
            "reward_id": parse_int(row[23]),
            "bp_exp": parse_int(row[24]),
        }
    wb.close()
    meta = {
        "task_pre_loc": task_pre_loc,
        "condition_pre_loc": condition_pre_loc,
        "patterns": patterns,
        "chains": chains,
        "tabs": tabs,
        "chain_rewards": chain_rewards,
    }
    return tasks, conds, meta


def has_translation(string_keys, lang_data, key, lang):
    if not key or key.lower() == "null":
        return True
    if key not in string_keys:
        return False
    h = string_keys[key]
    if lang not in lang_data:
        return False
    return h in lang_data[lang] and bool(lang_data[lang][h].strip())


def check_string_key(string_keys, lang_data, key):
    """Return list of (issue_type, msg, sev, table, extra)."""
    issues = []
    if not key or key.lower() == "null":
        issues.append(("StringKey为空", "StringKey未配置", "高", "Mission.xlsx/Task", key))
        return issues
    if key not in string_keys:
        issues.append(("StringKey未注册", "StringKey %s 不在 StringKeyTable.asset" % key, "高", "StringKeyTable.asset", key))
        return issues
    if not has_translation(string_keys, lang_data, key, "zh-Hans"):
        issues.append(("中文缺译", "StringKey %s 中文(zh-Hans)无译文或为空" % key, "高", "Lang/zh-Hans.asset", key))
    else:
        missing_langs = [lc for lc in FOREIGN_LANGS if not has_translation(string_keys, lang_data, key, lc)]
        if missing_langs:
            issues.append((
                "外语缺译",
                "StringKey %s 中文有译文，但 %s 缺译文（运行时可能fallback）" % (key, ",".join(missing_langs)),
                "中",
                "Lang/*.asset",
                "%s|%s" % (key, ",".join(missing_langs)),
            ))
    return issues


def collect_task_string_keys(task, conds, meta):
    keys = set()
    if task.get("title_key"):
        keys.add(task["title_key"])
    if task.get("context_key"):
        keys.add(task["context_key"])
    tid = task["task_id"]
    ploc = meta["task_pre_loc"].get(tid)
    if ploc:
        if ploc["title_key"]:
            keys.add(ploc["title_key"])
        if ploc["context_key"]:
            keys.add(ploc["context_key"])
    chain = meta["chains"].get(task.get("chain", ""))
    if chain:
        if chain.get("name_key"):
            keys.add(chain["name_key"])
        if chain.get("desc_key"):
            keys.add(chain["desc_key"])
        tab = chain.get("tab_id")
        if tab is not None and str(tab) in meta["tabs"]:
            keys.add(meta["tabs"][str(tab)])
    for cid in task.get("conds", []):
        c = conds.get(cid)
        if c and c.get("text_key"):
            keys.add(c["text_key"])
        cloc = meta["condition_pre_loc"].get(cid)
        if cloc:
            pt = cloc.get("pattern")
            if pt in meta["patterns"] and meta["patterns"][pt]:
                keys.add(meta["patterns"][pt])
            for t in cloc.get("text_list", []):
                if t.startswith("TXT_"):
                    keys.add(t)
    return keys


def validate_i18n(task, conds, meta, string_keys, lang_data, missing_lang_assets):
    issues = []
    tid = task["task_id"]

    if tid not in meta["task_pre_loc"]:
        issues.append((
            "多语言", "任务本地化映射缺失", "TaskID %s 在 TaskPreLoc 表无对应行" % tid,
            "高", "TaskPreLoc", "", "",
        ))
    else:
        ploc = meta["task_pre_loc"][tid]
        t = tasks_get(task)
        for field, ploc_key, task_key in [
            ("标题", ploc["title_key"], task.get("title_key")),
            ("描述", ploc["context_key"], task.get("context_key")),
        ]:
            if task_key and ploc_key and ploc_key != task_key:
                issues.append((
                    "多语言", "本地化映射不一致",
                    "TaskPreLoc与Task表%s StringKey不一致：PreLoc=%s Task=%s" % (field, ploc_key, task_key),
                    "中", "TaskPreLoc", ploc_key, "",
                ))

    for cid in task.get("conds", []):
        if cid not in meta["condition_pre_loc"]:
            issues.append((
                "多语言", "条件本地化映射缺失",
                "ConditionID %s 在 ConditionPreLoc 表无对应行" % cid,
                "高", "ConditionPreLoc", "", cid,
            ))
            continue
        cloc = meta["condition_pre_loc"][cid]
        pt = cloc.get("pattern")
        if pt not in meta["patterns"]:
            issues.append((
                "多语言", "句式模板缺失",
                "条件%s PatternType=%s 在 ConditionPatternStringkey 不存在" % (cid, pt),
                "高", "ConditionPatternStringkey", "", cid,
            ))
        elif meta["patterns"].get(pt):
            for itype, msg, sev, table, sk in check_string_key(string_keys, lang_data, meta["patterns"][pt]):
                issues.append(("多语言", itype, msg, sev, table, sk, cid))
        tl, cl = cloc.get("text_list", []), cloc.get("color_list", [])
        if len(tl) != len(cl):
            issues.append((
                "多语言", "颜色参数不匹配",
                "条件%s TextList(%d项)与ColorList(%d项)长度不一致" % (cid, len(tl), len(cl)),
                "中", "ConditionPreLoc", "", cid,
            ))
        for t in tl:
            if t.startswith("TXT_"):
                for itype, msg, sev, table, sk in check_string_key(string_keys, lang_data, t):
                    issues.append(("多语言", itype, msg, sev, table, sk, cid))

    for key in collect_task_string_keys(task, conds, meta):
        for itype, msg, sev, table, sk in check_string_key(string_keys, lang_data, key):
            issues.append(("多语言", itype, msg, sev, table, sk, ""))
    return issues


def tasks_get(task):
    return task


def validate_cross_ref(task, conds, meta, refs, all_task_ids):
    issues = []
    tid = task["task_id"]
    t = task

    if t.get("reward_id") and t["reward_id"] not in refs["reward"]:
        issues.append((
            "跨表引用", "RewardID无效",
            "RewardID=%s 在 Reward.xlsx 不存在" % t["reward_id"],
            "高", "Reward.xlsx", "", "",
        ))

    for label, did in [("接取对话", t.get("get_dialog")), ("领奖对话", t.get("claim_dialog"))]:
        if did and did not in refs["dialog"]:
            issues.append((
                "跨表引用", "DialogID无效",
                "%s TaskGetDialogID/TaskClaimedDialogID=%s 在 Dialog.xlsx 不存在" % (label, did),
                "中", "Dialog.xlsx", "", "",
            ))
        elif did and did in refs["dialog"] and did not in refs["dialog_has_content"]:
            issues.append((
                "跨表引用", "DialogContent缺失",
                "%s：DialogID=%s 在 Dialog.xlsx 存在，但 DialogContent 未找到可用内容" % (label, did),
                "中", "DialogContent", "", "",
            ))

    pre = t.get("pre_task", "")
    if pre and pre.lower() not in ("null", "none", ""):
        pre_id = str(parse_int(pre) or pre)
        if pre_id not in all_task_ids:
            issues.append((
                "跨表引用", "前置任务无效",
                "PreTaskID=%s 在 Task 表不存在" % pre,
                "高", "Task", "", "",
            ))

    chain = t.get("chain", "")
    if chain and chain not in meta["chains"]:
        issues.append((
            "跨表引用", "任务链无效",
            "TaskChain=%s 在 TaskChain 表不存在" % chain,
            "中", "TaskChain", "", "",
        ))
    elif chain and chain in meta["chains"]:
        tab_id = meta["chains"][chain].get("tab_id")
        if tab_id is not None and str(tab_id) not in meta["tabs"]:
            issues.append((
                "跨表引用", "任务页签无效",
                "TaskChain %s 的 TaskTabID=%s 在 TaskTab 不存在" % (chain, tab_id),
                "中", "TaskTab", "", "",
            ))

    if t.get("type") and 2000 <= int(t["type"]) <= 3000:
        if not t.get("bp_exp"):
            issues.append((
                "任务表", "BPExp缺失",
                "BP任务(TaskType=%s)未配置 BPExp" % t["type"],
                "中", "Task", "", "",
            ))

    unlock = parse_int(t.get("unlock"))
    if unlock and unlock != 0 and not t.get("unlock_param"):
        issues.append((
            "任务表", "解锁参数缺失",
            "UnlockCdt=%s 但 UnlockCdtValue 为空" % unlock,
            "中", "Task", "", "",
        ))

    for cid in t.get("conds", []):
        c = conds.get(cid)
        if not c:
            continue
        goto = c.get("goto")
        if goto and goto not in refs["jump_all"]:
            issues.append((
                "跨表引用", "跳转ID无效",
                "条件%s GotoViewPanel=%s 在 Jumptopanel.xlsx 不存在" % (cid, goto),
                "中", "Jumptopanel.xlsx", "", cid,
            ))
        elif goto and goto in refs["jump_disabled"]:
            issues.append((
                "跨表引用", "跳转可能被禁止",
                "条件%s GotoViewPanel=%s 在 Jumptopanel.xlsx 中 DisabledCondition=1（匹配时可能禁止跳转）" % (cid, goto),
                "中", "Jumptopanel.xlsx", "", cid,
            ))
        for pt, cmpv, pv in c.get("pre", []):
            ref = PRECDT_REF_TABLE.get(pt)
            if not ref:
                continue
            vid = parse_int(pv)
            if vid is None:
                continue
            table_name, _ = ref
            key = table_name.split(".")[0].lower()
            ref_key = {
                "monster": "monster", "item": "item", "class": "class",
                "rank": "rank", "collection": "collection", "level": "level",
                "makeitem": "makeitem", "chest": "chest", "merchant": "monster",
            }.get(key, key)
            pool = refs.get(ref_key, set())
            if pool and vid not in pool:
                issues.append((
                    "跨表引用", "PreCdt目标无效",
                    "条件%s PreCdt=%s(%s) 值=%s 在 %s 不存在" % (
                        cid, pt, PRECDT_NAME.get(pt, ""), pv, table_name),
                    "高", table_name, "", cid,
                ))
    return issues


def validate_task_rules(tid, t, conds):
    issues = []
    if not t:
        issues.append(("2.3规则", "任务缺失", "Mission.xlsx Task表找不到该任务ID", "高", "Task", "", ""))
        return issues
    if not t.get("conds"):
        issues.append(("任务表", "无条件", "任务未配置任何Condition ID", "中", "Task", "", ""))

    for cid in t.get("conds", []):
        c = conds.get(cid)
        if not c:
            issues.append(("2.3规则", "条件缺失", "Condition表找不到条件ID %s" % cid, "高", "Condition", "", cid))
            continue
        cdt = c["cdt"]
        if cdt in INGAME_CDT and not c["ingame_flag"]:
            issues.append((
                "2.3规则", "局内标记错误",
                "条件%s：CdtType=%s按2.3应为局内，但ShowInMatch=False" % (cid, cdt),
                "高", "Condition", "", cid,
            ))
        if cdt not in INGAME_CDT and c["ingame_flag"]:
            issues.append((
                "2.3规则", "局内标记错误",
                "条件%s：CdtType=%s按2.3应为局外，但ShowInMatch=True" % (cid, cdt),
                "中", "Condition", "", cid,
            ))

        family = CDT_PRECDT_FAMILY.get(cdt, set())
        pre_types_list = [pt for pt, _, _ in c["pre"]]
        pre_types = set(pre_types_list)
        if family and pre_types and not (pre_types & family) and not (pre_types & COMMON_PRECDT):
            issues.append((
                "2.3规则", "PreCdt与CdtType不匹配",
                "条件%s：CdtType=%s(%s) 的PreCdt %s 与类型不匹配" % (
                    cid, cdt, CDT_NAME.get(cdt, ""), ",".join(str(p) for p in pre_types)),
                "中", "Condition", "", cid,
            ))

        # 冗余检查：同一 Condition 槽位重复配了相同 PreCdt
        if len(pre_types_list) != len(pre_types):
            issues.append((
                "2.3规则", "重复PreCdt",
                "条件%s：存在重复的PreCdt配置(%s)" % (cid, ",".join(str(p) for p in sorted(pre_types))),
                "中", "Condition", "", cid,
            ))

        for pt, cmpv, pv in c["pre"]:
            if pt not in VALID_PRECDT:
                issues.append((
                    "2.3规则", "PreCdt枚举无效",
                    "条件%s：PreCdt=%s 不在有效枚举范围" % (cid, pt),
                    "高", "Condition", "", cid,
                ))
            if pt in DEPRECATED_PRECDT:
                issues.append((
                    "2.3规则", "废弃PreCdt",
                    "条件%s：%s" % (cid, DEPRECATED_PRECDT[pt]),
                    "高", "Condition", "", cid,
                ))
            if cdt == 15 and pt == 12 and cmpv != 1:
                issues.append((
                    "2.3规则",
                    "比较符错误",
                    "条件%s：段位PreCdt应使用>=（PreCdt=12），当前比较符=%s" % (cid, CMP_NAME.get(cmpv, str(cmpv))),
                    "中", "Condition", "", cid,
                ))
            if cdt == 16 and pt == 13 and cmpv != 1:
                issues.append((
                    "2.3规则",
                    "比较符错误",
                    "条件%s：收藏价值PreCdt应使用>=（PreCdt=13），当前比较符=%s" % (cid, CMP_NAME.get(cmpv, str(cmpv))),
                    "中", "Condition", "", cid,
                ))
            if cdt == 26 and pt != 42:
                issues.append((
                    "2.3规则", "2.3职业解锁",
                    "条件%s：应使用PreCdt=42(UNLOCKED_CLASS_GROUP_ID)，当前=%s" % (cid, pt),
                    "高", "Condition", "", cid,
                ))
            if cdt == 26 and len(c["pre"]) > 1:
                issues.append(("2.3规则", "2.3职业解锁", "条件%s：职业解锁PreCdt只能配单个" % cid, "中", "Condition", "", cid))
            if cdt == 34 and pt == 41 and len(c["pre"]) > 1:
                issues.append(("2.3规则", "套装完成", "条件%s：COLLECTION_ID只能配单个" % cid, "中", "Condition", "", cid))
            if pt == 28 and cmpv not in (None, 0):
                issues.append(("2.3规则", "怪物Tag", "条件%s：MONSTER_TAG仅可配等于(=)" % cid, "中", "Condition", "", cid))
            if cdt == 17 and pt == 32:
                issues.append((
                    "2.3规则", "2.3打造",
                    "条件%s：不应再使用PreCdt=32，应改用PreCdt=36(配方子类型)" % cid,
                    "高", "Condition", "", cid,
                ))
            if pt in {9, 39, 43} and cdt == 26:
                issues.append((
                    "2.3规则", "精通/职业组混用",
                    "条件%s：CdtType=26(职业组解锁)不应使用精通类PreCdt=%s" % (cid, pt),
                    "高", "Condition", "", cid,
                ))
            if pv in (None, "", "null") and pt not in {14, 16, 48}:
                issues.append((
                    "2.3规则", "PreCdtValue为空",
                    "条件%s：PreCdt=%s 需要配置值但为空" % (cid, pt),
                    "中", "Condition", "", cid,
                ))

        if c.get("text_key"):
            for itype, msg, sev, table, sk in []:
                pass  # i18n handled separately

    return issues


def fmt_pre(pt, cmpv, pv):
    name = PRECDT_NAME.get(pt, "PreCdt" + str(pt))
    cmp_s = CMP_NAME.get(cmpv if cmpv is not None else 0, str(cmpv))
    return "%s %s %s" % (name, cmp_s, pv)


def classify_preconditions(pre_list):
    ingame_pre, outgame_pre, setup_pre = [], [], []
    for pt, cmpv, pv in pre_list:
        s = fmt_pre(pt, cmpv, pv)
        if pt in {15, 19, 25, 34, 49, 31, 20, 21, 22, 23, 24, 45, 9, 52, 53, 43}:
            setup_pre.append("进局前准备：" + s)
        if pt in {1, 2, 10, 16, 17, 18, 28, 29, 30, 26, 27, 3, 8, 11, 47, 48}:
            ingame_pre.append("局内条件：" + s)
        if pt in {4, 5, 6, 12, 13, 33, 36, 39, 40, 41, 42, 44, 50, 51}:
            outgame_pre.append("局外/提交侧：" + s)
        if pt in {14}:
            setup_pre.append("进局前准备：需成功撤离才算（IS_WIN）")
    return ingame_pre, outgame_pre, setup_pre


def build_condition_summaries(cids, conds):
    ingame_checks, outgame_checks, setup_checks = [], [], []
    for cid in cids:
        c = conds.get(cid)
        if not c:
            ingame_checks.append("【缺配置】条件%s在Condition表不存在" % cid)
            continue
        cdt = c["cdt"]
        cname = CDT_NAME.get(cdt, "CdtType" + str(cdt))
        target = "目标值=%s" % c["cdt_val"] if c["cdt_val"] not in (None, "", "null") else ""
        base = "条件%s【%s】%s" % (cid, cname, target)
        if c["remark"]:
            base += "（%s）" % c["remark"]
        if cdt in INGAME_CDT or c["ingame_flag"]:
            ingame_checks.append(base)
        else:
            outgame_checks.append(base)
        ip, op, sp = classify_preconditions(c["pre"])
        ingame_checks.extend(ip)
        outgame_checks.extend(op)
        setup_checks.extend(sp)
    return ingame_checks, outgame_checks, setup_checks


def build_steps(task, conds):
    ing, out, setup = build_condition_summaries(task["conds"], conds)
    has_ing = bool(ing)
    has_out = bool(out) or bool(setup)

    accept_out = [
        "任务列表能看到本任务，标题和描述正确",
        "任务ID=" + task["task_id"],
    ]
    if task.get("pre_task") not in ("", "null", "None"):
        accept_out.append("前置任务 " + task["pre_task"] + " 已完成")
    if task.get("unlock") not in ("", "null", "None"):
        accept_out.append("解锁条件=" + task["unlock"] + "，参数=" + task["unlock_param"])

    accept_action = (
        "进大厅打开【任务】，确认任务已在列表（自动接取）"
        if task["auto_accept"]
        else "进大厅打开【任务】，手动点击【接取】"
    )

    steps = [{
        "phase": "接取", "location": "局外", "action": accept_action,
        "ingame_cp": "", "outgame_cp": "；".join(accept_out),
        "setup_cp": "；".join(setup) if setup else "",
        "expected": "任务变为【进行中】，目标描述和PreCondition提示正常",
    }]

    if has_ing and has_out:
        loc, action = "局外+局内", "先按局外检查点完成大厅侧目标，再开一局完成局内目标。任务说明：" + task["desc"]
    elif has_ing:
        loc, action = "局内", "开一局对局，在局内完成目标。任务说明：" + task["desc"]
    else:
        loc, action = "局外", "在大厅完成目标，无需进局内。任务说明：" + task["desc"]

    steps.append({
        "phase": "执行", "location": loc, "action": action,
        "ingame_cp": "；".join(ing), "outgame_cp": "；".join(out),
        "setup_cp": "；".join(setup) if setup else "",
        "expected": "进度条/计数与Condition配置一致，局内任务在局内HUD可见",
    })
    steps.append({
        "phase": "完成", "location": "局外", "action": "回大厅【任务】界面领取奖励",
        "ingame_cp": "", "outgame_cp": "奖励到账；任务变【已完成】；下一条前置任务解锁",
        "setup_cp": "",
        "expected": "领奖成功无报错，可继续下一条顺序号",
    })
    return steps


def load_all_tasks(tasks):
    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    tpl = next(t for t in cfg["savedTemplates"] if t["title"] == "老任务")
    all_tasks = []
    for row in tpl["aoa"][1:]:
        tid = str(parse_int(row[0]))
        t = tasks.get(tid, {})
        all_tasks.append({
            "category": "老任务模板",
            "task_id": tid,
            "name": str(row[1]).strip() or t.get("name", ""),
            "desc": str(row[2]).strip() or t.get("desc", ""),
            "chain": str(parse_int(row[7]) or t.get("chain", "")),
            "chain_name": CHAIN_NAMES.get(str(parse_int(row[7])), "任务链" + str(row[7])),
            "auto_accept": str(row[14]).strip() not in ("", "0", "null", "False"),
            "pre_task": str(row[17]).strip(),
            "unlock": str(row[15]).strip(),
            "unlock_param": str(row[16]).strip(),
            "conds": t.get("conds") or split_semicolon_list(row[21]),
        })
    all_tasks.sort(key=lambda x: (int(x["chain"] or 0), int(x["task_id"])))

    for d in sorted([v for v in tasks.values() if str(v.get("type")) == "6"], key=lambda x: int(x["task_id"])):
        all_tasks.append({
            "category": "日常首胜", "task_id": d["task_id"], "name": "日常首胜",
            "desc": d.get("desc") or "当日首场对局胜利", "chain": "", "chain_name": "日常首胜",
            "auto_accept": d["auto_accept"], "pre_task": d["pre_task"], "unlock": d["unlock"],
            "unlock_param": d["unlock_param"], "conds": d["conds"],
        })

    bp = [v for v in tasks.values() if v.get("type") and 2000 <= int(v["type"]) <= 3000]
    bp.sort(key=lambda d: (d["type"], int(str(d.get("chain") or 0) or 0), int(d["task_id"])))
    for d in bp:
        all_tasks.append({
            "category": "当期BP任务", "task_id": d["task_id"], "name": d.get("name") or ("BP-" + d["task_id"]),
            "desc": d.get("desc") or "", "chain": str(d["type"]), "chain_name": "BP类型" + str(d["type"]),
            "auto_accept": d["auto_accept"], "pre_task": d["pre_task"], "unlock": d["unlock"],
            "unlock_param": d["unlock_param"], "conds": d["conds"],
        })
    return all_tasks


def merge_task_for_validation(task, tasks):
    t = dict(task)
    if task["task_id"] in tasks:
        t.update({k: v for k, v in tasks[task["task_id"]].items() if k not in t or not t.get(k)})
    return t


def dedupe_issues(issue_list):
    seen = set()
    out = []
    for row in issue_list:
        key = tuple(row)
        if key not in seen:
            seen.add(key)
            out.append(row)
    return out


def build_check_guide():
    return [
        ["配置检查项说明"],
        ["生成时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["问题分类", "检查项", "规则来源", "严重度说明"],
        ["多语言", "TaskPreLoc / ConditionPreLoc 映射是否存在", "Mission.xlsx + 客户端展示逻辑", "缺失=高"],
        ["多语言", "StringKey 是否在 StringKeyTable 注册", "StringKeyTable.asset", "未注册=高"],
        ["多语言", "zh-Hans 是否有非空译文", "Lang/zh-Hans.asset", "中文空=高"],
        ["多语言", "en/pt/vi/th/id/es/ms 是否缺译（相对中文）", "Lang/*.asset", "外语缺译=中（可能fallback）"],
        ["多语言", "ConditionPreLoc TextList/ColorList 长度对齐", "ConditionPreLoc", "不一致=中"],
        ["多语言", "ja 语种资源文件是否存在", "LocalizationBrowserWindow", "缺失=中"],
        ["任务表", "Conditions / RewardID / BPExp / 解锁参数", "Mission.xlsx Task", "按项标注"],
        ["2.3规则", "ShowInMatch vs CdtType 局内集合", "2.3 Condition文档", "不一致=高/中"],
        ["2.3规则", "废弃 PreCdt 7/32；职业解锁 PreCdt=42", "2.3 Condition文档", "高"],
        ["2.3规则", "PreCdt 与 CdtType 家族匹配", "2.3 + 配置说明", "中"],
        ["跨表引用", "Reward / Dialog / JumpToPanel / PreTask / TaskChain", "各 Excel 表", "高/中"],
        ["跨表引用", "PreCdt 目标 ID（怪物/物品/职业等）", "Monster/Item/Class 等", "高"],
        [],
        ["优先级", "2.3 PDF > 0.9 plus > Mission.xlsx 配置说明"],
        ["筛选建议", "在「配置问题清单」按「问题分类」列筛选后分派策划/本地化/数值"],
    ]


def write_sheet(wb, name, rows, widths):
    ws = wb.create_sheet(name)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = wrap
            if r == 1:
                cell.fill = fill
                cell.font = font
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def main():
    print("loading mission tables...")
    tasks, conds, meta = load_mission_tables()
    print("loading i18n...")
    string_keys, lang_data, missing_lang_assets = load_i18n()
    print("loading ref tables...")
    refs = load_ref_tables()
    all_task_ids = set(tasks.keys())
    all_tasks = load_all_tasks(tasks)

    headers = [
        "顺序号", "任务分类", "任务链", "任务链名称", "任务ID", "任务名称", "任务描述",
        "阶段", "局内/局外", "操作步骤（大白话）", "局内检查点", "局外检查点",
        "进局前准备(PreCondition)", "预期结果", "备注",
    ]
    detail = [headers]
    issue_header = [
        "任务分类", "任务链", "任务ID", "任务名称", "问题分类", "问题类型",
        "问题说明", "严重级别", "配置表", "涉及StringKey/语种", "涉及条件ID",
    ]
    issue_rows = [issue_header]
    issue_stats = {}

    # 语种资源缺失属于“全局配置问题”，只输出一次
    for lc in missing_lang_assets:
        issue_stats["多语言"] = issue_stats.get("多语言", 0) + 1
        issue_rows.append([
            "全局", "", "", "",
            "多语言", "语种资源缺失",
            "代码支持语种 %s 但 Lang/%s.asset 不存在" % (lc, lc),
            "中", "Lang/", lc, "",
        ])

    seq = 1
    for task in all_tasks:
        t = merge_task_for_validation(task, tasks)
        task["conds"] = t.get("conds", [])

        all_issues = []
        all_issues.extend(validate_task_rules(task["task_id"], t if task["task_id"] in tasks else None, conds))
        all_issues.extend(validate_i18n(t, conds, meta, string_keys, lang_data, missing_lang_assets))
        all_issues.extend(validate_cross_ref(t, conds, meta, refs, all_task_ids))

        for row in dedupe_issues(all_issues):
            cat = row[0]
            issue_stats[cat] = issue_stats.get(cat, 0) + 1
            issue_rows.append([
                task["category"], task["chain"], task["task_id"], task["name"],
            ] + list(row))

        note = []
        if task.get("pre_task") not in ("", "null", "None"):
            note.append("前置=" + task["pre_task"])
        note.append("自动接取=" + ("是" if task["auto_accept"] else "否"))
        for step in build_steps(task, conds):
            detail.append([
                seq, task["category"], task["chain"], task["chain_name"], task["task_id"],
                task["name"], task["desc"], step["phase"], step["location"], step["action"],
                step["ingame_cp"], step["outgame_cp"], step["setup_cp"], step["expected"], "；".join(note),
            ])
            seq += 1

    guide = [
        ["任务全量用例（2.3规则 + 多语言 + 跨表引用全量检查）"],
        ["生成时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["参考文档"],
        ["优先级", "【2.3】任务Condition汇总（职业精通/职业组/废弃项以2.3为准）"],
        ["辅助", "【0.9 plus】任务Condition汇总"],
        ["数据源", "Mission.xlsx + easydone「老任务」模板 + Lang/*.asset"],
        [],
        ["怎么用"],
        ["1", "按【用例明细】顺序号执行：接取→执行→完成"],
        ["2", "局内检查点：必须进对局验证；局外检查点：大厅/任务界面验证"],
        ["3", "进局前准备：按PreCondition在大厅配好队、职业、装备、携带物再开局"],
        ["4", "【配置问题清单】按问题分类筛选：多语言/2.3规则/跨表引用/任务表"],
        ["5", "【配置检查项说明】查看完整检查规则与严重度定义"],
        [],
        ["问题统计（按问题分类）"],
    ] + [[k, str(v)] for k, v in sorted(issue_stats.items(), key=lambda x: -x[1])]

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "使用说明", guide, {1: 28, 2: 72})
    write_sheet(wb, "用例明细", detail, {
        1: 8, 2: 12, 3: 8, 4: 16, 5: 10, 6: 14, 7: 30, 8: 8, 9: 10, 10: 36,
        11: 40, 12: 40, 13: 36, 14: 24, 15: 18,
    })
    write_sheet(wb, "配置问题清单", issue_rows, {
        1: 12, 2: 8, 3: 10, 4: 16, 5: 12, 6: 14, 7: 48, 8: 10, 9: 18, 10: 28, 11: 14,
    })
    write_sheet(wb, "配置检查项说明", build_check_guide(), {1: 14, 2: 36, 3: 28, 4: 16})
    wb.save(OUT_PATH)
    print("saved", OUT_PATH)
    print("tasks", len(all_tasks), "steps", len(detail) - 1, "issues", len(issue_rows) - 1)
    print("issue_stats", issue_stats)


if __name__ == "__main__":
    main()
